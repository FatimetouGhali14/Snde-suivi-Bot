"""
Service de fusion des rapports Excel du jour en un seul fichier.

Récupère les fichiers Telegram des rapports reçus aujourd'hui et les
combine dans un classeur unique avec un onglet par direction.
"""
from __future__ import annotations

import io
import logging
import tempfile
from datetime import date as Date

import openpyxl
from telegram import Bot

from src.db.pool import get_connexion

logger = logging.getLogger(__name__)


def recuperer_rapports_avec_fichier_du_jour() -> list[dict]:
    """Retourne les rapports actifs du jour avec leur file_id."""
    conn = get_connexion()
    cursor = conn.cursor()
    cursor.execute(
        """
        SELECT 
            d.direction_code,
            d.nom_complet,
            r.telegram_file_id,
            r.nom_fichier,
            r.heure_envoi
        FROM rapports_quotidiens r
        JOIN directeurs d ON r.directeur_uuid = d.id
        WHERE r.date_rapport = CURRENT_DATE
          AND r.telegram_file_id IS NOT NULL
          AND COALESCE(r.est_actif, TRUE) = TRUE
        ORDER BY r.heure_envoi
        """
    )
    rows = cursor.fetchall()
    return [
        {
            "direction_code": r[0],
            "nom_complet": r[1],
            "file_id": r[2],
            "nom_fichier": r[3],
            "heure_envoi": r[4],
        }
        for r in rows
    ]


async def construire_fichier_consolide(bot: Bot) -> bytes | None:
    """
    Télécharge tous les fichiers Excel du jour depuis Telegram
    et les fusionne en un classeur unique.
    
    Returns:
        Les bytes du fichier consolidé, ou None si aucun rapport.
    """
    rapports = recuperer_rapports_avec_fichier_du_jour()
    
    if not rapports:
        logger.info("Aucun rapport avec fichier pour aujourd'hui")
        return None
    
    # Classeur de destination
    wb_dest = openpyxl.Workbook()
    # Supprimer l'onglet par défaut "Sheet"
    wb_dest.remove(wb_dest.active)
    
    for rapport in rapports:
        code = rapport["direction_code"]
        file_id = rapport["file_id"]
        
        try:
            # Télécharger le fichier Telegram dans un fichier temporaire
            fichier_tg = await bot.get_file(file_id)
            with tempfile.NamedTemporaryFile(
                suffix=".xlsx", delete=False
            ) as tmp:
                chemin_tmp = tmp.name
            await fichier_tg.download_to_drive(chemin_tmp)
            
            # Ouvrir le classeur source
            wb_src = openpyxl.load_workbook(chemin_tmp, data_only=False)
            
            # Trouver l'onglet de la direction
            onglet_source = None
            if code in wb_src.sheetnames:
                onglet_source = wb_src[code]
            elif len(wb_src.sheetnames) == 1:
                onglet_source = wb_src[wb_src.sheetnames[0]]
            
            if onglet_source is None:
                logger.warning("Onglet %s introuvable dans le fichier", code)
                wb_src.close()
                continue
            
            # Copier l'onglet dans le classeur de destination
            ws_dest = wb_dest.create_sheet(title=code)
            for row in onglet_source.iter_rows(values_only=False):
                for cell in row:
                    new_cell = ws_dest.cell(
                        row=cell.row,
                        column=cell.column,
                        value=cell.value,
                    )
            
            # Copier les largeurs de colonnes
            for col_letter, col_dim in onglet_source.column_dimensions.items():
                if col_dim.width:
                    ws_dest.column_dimensions[col_letter].width = col_dim.width
            
            wb_src.close()
            
            import os
            os.unlink(chemin_tmp)
            
            logger.info("Onglet %s ajoute au fichier consolide", code)
            
        except Exception as e:
            logger.error("Erreur fusion %s : %s", code, e)
            continue
    
    if len(wb_dest.sheetnames) == 0:
        logger.warning("Aucun onglet n'a pu etre fusionne")
        return None
    
    # Sauvegarder en bytes
    buffer = io.BytesIO()
    wb_dest.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()