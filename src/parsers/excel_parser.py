# src/parsers/excel_parser.py
# Lecture du fichier Excel SNDE — rapport journalier multi-onglets

import logging
from dataclasses import dataclass, field
from pathlib import Path
import openpyxl
from copy import copy
import io

logger = logging.getLogger(__name__)

# Correspondance onglet → direction
ONGLETS_DIRECTIONS = {
    "PROD"  : "Direction de la Production",
    "DIST"  : "Direction de la Distribution",
    "MAINT" : "Direction de la Maintenance",
    "ETN"   : "Direction des Etudes & Travaux Neufs",
    "STOCK" : "Direction des Stocks",
    "ACHAT" : "Direction des Achats",
    "COM"   : "Direction Commerciale",
    "INFO"  : "Direction de l'Informatique",
    "QUAL"  : "Direction de la Qualite de l'Eau",
    "FIN"   : "Direction Financiere & Comptable",
    "URG"   : "Direction des Urgences",
    "AUDIT" : "Direction de l'Audit & Controle",
}


@dataclass
class KPI:
    """Un indicateur clé de performance."""
    numero       : int
    indicateur   : str
    unite        : str
    valeur_j1    : str = ""
    valeur_jour  : str = ""
    objectif     : str = ""
    observations : str = ""
    statut       : str = ""


@dataclass
class RapportDirection:
    """Rapport complet d'une direction."""
    direction_code   : str
    direction_nom    : str
    date_rapport     : str = ""
    directeur        : str = ""
    kpis             : list[KPI] = field(default_factory=list)
    points_bloquants : list[str] = field(default_factory=list)
    actions_demain   : list[str] = field(default_factory=list)
    instructions_dg  : list[str] = field(default_factory=list)
    heure_envoi      : str = ""
    complet          : bool = False


def lire_fichier_complet(chemin: str | Path) -> dict[str, RapportDirection]:
    """
    Lit le fichier Excel complet et retourne un dict
    avec les données de chaque direction.
    """
    chemin = Path(chemin)
    if not chemin.exists():
        raise FileNotFoundError(f"Fichier introuvable : {chemin}")

    wb = openpyxl.load_workbook(chemin, data_only=True)
    rapports = {}

    for code_onglet, nom_direction in ONGLETS_DIRECTIONS.items():
        if code_onglet not in wb.sheetnames:
            logger.warning(f"Onglet manquant : {code_onglet}")
            continue

        ws = wb[code_onglet]
        rapport = extraire_rapport_direction(ws, code_onglet, nom_direction)
        rapports[code_onglet] = rapport
        logger.info(f"Onglet lu : {code_onglet} — {len(rapport.kpis)} KPIs")

    wb.close()
    return rapports


def extraire_rapport_direction(ws, code: str, nom: str) -> RapportDirection:
    """Extrait les données d'un onglet direction."""
    rapport = RapportDirection(
        direction_code = code,
        direction_nom  = nom,
    )

    # Lire toutes les valeurs dans une liste
    rows = list(ws.iter_rows(values_only=True))

    # Ligne 2 : Date et Directeur (colonnes E et B)
    if len(rows) >= 2:
        rapport.date_rapport = str(rows[1][5]) if rows[1][5] else ""
        rapport.directeur    = str(rows[2][1]) if len(rows) > 2 and rows[2][1] else ""

    # Lignes KPIs : lignes 8 à 15 (index 7 à 14)
    kpis = []
    for i in range(7, 15):
        if i >= len(rows):
            break
        row = rows[i]
        if not row[0]:
            continue
        try:
            kpi = KPI(
                numero       = int(row[0]) if row[0] else i - 6,
                indicateur   = str(row[1]) if row[1] else "",
                unite        = str(row[2]) if row[2] else "",
                valeur_j1    = str(row[3]) if row[3] else "",
                valeur_jour  = str(row[4]) if row[4] else "",
                objectif     = str(row[5]) if row[5] else "",
                observations = str(row[6]) if row[6] else "",
                statut       = str(row[7]) if row[7] else "",
            )
            kpis.append(kpi)
        except Exception:
            continue

    rapport.kpis = kpis

    # Vérifier si le rapport est rempli
    valeurs_remplies = [
        k.valeur_jour for k in kpis
        if k.valeur_jour and k.valeur_jour != "None"
    ]
    rapport.complet = len(valeurs_remplies) >= 3

    return rapport


def extraire_onglet_vers_bytes(chemin: str | Path, code_onglet: str) -> bytes:
    """
    Extrait un onglet du fichier Excel et retourne
    les bytes d'un nouveau fichier Excel avec cet onglet uniquement.
    """
    chemin = Path(chemin)
    wb_source = openpyxl.load_workbook(chemin)

    if code_onglet not in wb_source.sheetnames:
        raise ValueError(f"Onglet {code_onglet} introuvable")

    # Créer un nouveau classeur avec cet onglet seulement
    wb_dest = openpyxl.Workbook()
    ws_dest = wb_dest.active
    ws_source = wb_source[code_onglet]
    ws_dest.title = code_onglet

    # Copier toutes les cellules
    for row in ws_source.iter_rows():
        for cell in row:
            new_cell = ws_dest.cell(
                row    = cell.row,
                column = cell.column,
                value  = cell.value
            )

    # Copier les dimensions des colonnes
    for col_letter, col_dim in ws_source.column_dimensions.items():
        ws_dest.column_dimensions[col_letter].width = col_dim.width

    # Sauvegarder en bytes
    buffer = io.BytesIO()
    wb_dest.save(buffer)
    buffer.seek(0)

    wb_source.close()
    return buffer.getvalue()


def verifier_kpis_remplis(rapport: RapportDirection) -> dict:
    """
    Vérifie quels KPIs ont été remplis.
    Retourne un résumé.
    """
    total = len(rapport.kpis)
    remplis = sum(
        1 for k in rapport.kpis
        if k.valeur_jour and k.valeur_jour != "None" and k.valeur_jour != ""
    )
    return {
        "total"   : total,
        "remplis" : remplis,
        "manquants": total - remplis,
        "complet" : remplis >= total * 0.75,
    }