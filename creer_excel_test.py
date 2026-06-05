# creer_excel_test.py
# Crée un fichier Excel de test pour le DG

import openpyxl
from datetime import date, timedelta

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Instructions DG"

# En-têtes
ws["A1"] = "Direction"
ws["B1"] = "Action"
ws["C1"] = "Délai"
ws["D1"] = "Priorité"

# Données de test
instructions = [
    ("DAF",            "Préparer le bilan financier du mois",         date.today() + timedelta(days=3),  "urgent"),
    ("DRH",            "Soumettre le rapport des effectifs",           date.today() + timedelta(days=5),  "normal"),
    ("DT",             "Planifier la maintenance des équipements",     date.today() + timedelta(days=7),  "normal"),
    ("DC",             "Préparer le rapport commercial mensuel",       date.today() + timedelta(days=4),  "urgent"),
    ("DInformatique",  "Mettre à jour les systèmes de sécurité",      date.today() + timedelta(days=2),  "urgent"),
]

for row in instructions:
    ws.append(row)

wb.save("test_instructions.xlsx")
print("Fichier test_instructions.xlsx créé avec succès !")